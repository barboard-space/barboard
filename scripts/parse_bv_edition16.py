#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""第十六届（Barvision Chongqing 2026）赛果解析 → regular-16.json 的 matches。

源：C:\\Users\\William Liu\\Downloads\\Barvision Chongqing 2026 Results.xlsx（用户提供的最终版）

⭐ 本届与 2019–2025 的根本差异（务必先读）
--------------------------------------------------
**每位投票人同时投「评委票」和「观众票」**——不再是 2019–2025 的「选送者=评委 / 其余=观众」两拨人。
故 votes.voters[] 里同一个人会出现两条记录（type=jury 一条、type=tele 一条），
渲染器对 tele_mode='votes' 的场次本来就把计分板拆成 Jury / Tele 两张表，互不干扰。

三场（SF1 / SF2 / GF）的 sheet 都是「上下两块同构表」：
  · 上块 = 评委逐票（Top 10 给 12/10/8/7/6/5/4/3/2/1，每人合计 58）
  · 下块 = 观众逐票原始票数（每人 20 票自由分配）
  · 两块共用同一组投票人列；某人某块全空 = 该块未投
  · 尾部三列：上块 = Jury(评委分) / Tele(观众分,四舍五入) / Total
              下块 = 总票数 / 得票率 / 观众分(未四舍五入)
观众分换算：观众分 = 票数 × (本场评委总分 ÷ 本场总票数)，三场系数约 2.9 / 2.7389 / 2.9。
⚠ **tele_vote / score 直接取表里已发布的整数值，不自己重算**——源表的四舍五入有若干处
（SF2 三处 x.50~x.51 向下、GF 一处 x.5 向下）不是任何统一规则能复现的浮点/人工痕迹，
自己算会让 6 首歌的总分与官方公布值差 1 分。逐票和只用来校验 Jury 列。

⚠ **选送者未参与本场投票 → 该曲总分按 50% 计**（田妈 SF1《The Wild》75→38、
杰妈 SF2《Feel The Rush》144→72）。沿用历史折算写法：jury_vote/tele_vote 存折算前原始分、
score 存折算后，match.note 说明。

另两场（Wildcard Round 海选突围赛 / Second Chance 外卡突围赛）是**认可票**（每人 1 票、至多 3 票）：
  · 单块表，score = 认可票数；Second Chance 另有 Penalty（未投票者扣 1 分）→ score = Pts（扣分后）
  · 名次直接取官方「名次」列（含并列），故这两场在 recompute_bv_ranks 里靠 format='approval' 跳过重算

用法：python scripts/parse_bv_edition16.py [--check]
      --check 只打印校验报告，不写文件
"""
import io
import json
import os
import sys

from openpyxl import load_workbook

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SRC = r"C:\Users\William Liu\Downloads\Barvision Chongqing 2026 Results.xlsx"
OUT = os.path.join(BASE, "data", "barvision", "barvision-2026", "regular-16.json")
CHECK_ONLY = "--check" in sys.argv

# 简称 → 规范昵称（联合选送/联合投票用斜杠串，下游按 / 拆分计入两人）
ABBR = {
    "锴": "锴妈", "萌": "萌妈", "波": "波妈", "A": "A妈", "团": "团妈", "奶": "奶妈",
    "包": "包妈", "麦": "麦妈", "韩": "韩妈", "叉": "叉妈", "柠": "柠妈", "白": "白妈",
    "吃": "吃妈", "猴": "猴妈", "田": "田妈", "邓": "邓妈", "城": "城妈", "松": "松妈",
    "XX": "XX妈", "K": "K妈", "P": "P妈", "星": "星妈", "雨": "雨妈", "笃": "笃妈",
    "威": "威妈", "羊": "羊妈", "野": "野妈", "米": "米妈", "嘟": "嘟妈", "海": "海妈",
    "柴": "柴妈", "泰": "泰妈", "汞": "汞妈", "T": "T妈", "X": "X妈", "杰": "杰妈",
    "文": "文妈", "瑞": "瑞妈", "风": "风妈", "圈": "圈妈", "L": "L妈",
    "S鸽": "S妈/鸽妈", "狼芬": "狼妈/芬妈",
}

# 昵称 → space_id（联合串不给 id，member_id=null）
MEMBER_ID = {
    "X妈": 195, "松妈": 105, "XX妈": 88, "韩妈": 67, "白妈": 77, "波妈": 101,
    "雨妈": 17, "笃妈": 33, "柠妈": 19, "A妈": 132, "海妈": 119, "邓妈": 31,
    "猴妈": 18, "羊妈": 100, "星妈": 45, "包妈": 20, "威妈": 7, "城妈": 120,
    "吃妈": 126, "团妈": 68, "S妈": 27, "鸽妈": 160, "柴妈": 51, "泰妈": 131,
    "汞妈": 103, "奶妈": 102, "叉妈": 13, "文妈": 116, "P妈": 104, "狼妈": 113,
    "芬妈": 110, "K妈": 11, "野妈": 196, "萌妈": 125, "杰妈": 156, "麦妈": 130,
    "T妈": 115, "锴妈": 12, "嘟妈": 190, "田妈": 295, "米妈": 256, "瑞妈": 135,
    "风妈": 154, "圈妈": 108, "L妈": 107,
}

HANDLE = {
    "X妈": "没有XX不科学", "松妈": "Squirrel松鼠", "XX妈": "xjebs", "韩妈": "SouthkoreaBall",
    "白妈": "白哼唧", "波妈": "微波子", "雨妈": "Lee翼雨", "笃妈": "吃辣喝香",
    "柠妈": "LemonSheeran", "A妈": "AloneAlien_", "海妈": "Ocean", "邓妈": "DuncanLee",
    "猴妈": "Voiiz", "羊妈": "SeafishYANG", "星妈": "白矮星爱吃东西", "包妈": "Jeremy_BAg",
    "威妈": "williw_", "城妈": "MarekTotti", "吃妈": "能吃的莉莉丝", "团妈": "不动团子",
    "S妈": "哈哈哈时光机", "鸽妈": "EUGONDINE_", "柴妈": "是柴柴呀", "泰妈": "泰坦crazy",
    "汞妈": "Hg", "奶妈": "奶儿少爷", "叉妈": "PRONGS", "文妈": "SummertimeDawns",
    "P妈": "PROKING", "狼妈": "Aaowu_", "芬妈": "Stephen", "K妈": "SoyMarvin",
    "野妈": "Tye", "萌妈": "绿荫夏语", "杰妈": "jh201013", "麦妈": "Tandiny",
    "T妈": "iTAP_II", "锴妈": "WillKris锴", "嘟妈": "DDDDDope", "田妈": "FREEMAN",
    "米妈": "LUKY", "瑞妈": "Jefferyat19", "风妈": "Durden", "圈妈": "TimCircles6",
    "L妈": "Lone",
}

MATCH_META = [
    ("Semi-Final 1", "SF1", "半决赛一"),
    ("Semi-Final 2", "SF2", "半决赛二"),
    ("Wildcard Round", "WC", "海选突围赛"),
    ("Second Chance", "SC", "外卡突围赛"),
    ("Grand Final", "GF", "决赛"),
]


def num(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def nick(abbr):
    a = str(abbr).strip()
    if a not in ABBR:
        raise KeyError("未知简称：%r" % a)
    return ABBR[a]


def mid(nickname):
    return None if "/" in nickname else MEMBER_ID[nickname]


def load_signup_meta():
    """语种 / 流派（人工核对过的权威值，源表没有这两列）→ 按歌手名索引。
    首次解析取自 signups（报名名单）；signups 在转正式届结构时已删除，
    之后重跑改从已生成的 matches 里取，保证脚本幂等可重跑。"""
    d = json.load(open(OUT, encoding="utf-8"))
    meta = {}
    for s in d.get("signups", []):
        meta[s["artist"].strip()] = {"language": s.get("language"), "genre": s.get("genre")}
    for m in d.get("matches", []):
        for e in m.get("entries", []):
            k = (e.get("artist") or "").strip()
            if k and k not in meta and e.get("language"):
                meta[k] = {"language": e.get("language"), "genre": e.get("genre")}
    return meta, d


def read_sheet(wb, sheet):
    rows = [r for r in wb[sheet].iter_rows(values_only=True)]
    hdr = rows[0]
    return hdr, rows


def split_blocks(rows):
    """上下两块同构表 → (block1, block2)；第二块表头行以 col0=='序号' 识别。"""
    h2 = [i for i, r in enumerate(rows) if i > 0 and r[0] == "序号"]
    if not h2:
        return [r for r in rows[1:] if num(r[0]) is not None], []
    cut = h2[0]
    b1 = [r for r in rows[1:cut] if num(r[0]) is not None]
    b2 = [r for r in rows[cut + 1:] if num(r[0]) is not None]
    return b1, b2


def build_dual(wb, sheet, code, venue, meta, report):
    """SF1 / SF2 / GF：评委块 + 观众块 → 一个 match。"""
    hdr, rows = read_sheet(wb, sheet)
    iJ, iT, iTot = hdr.index("Jury"), hdr.index("Tele"), hdr.index("Total")
    nmeta = 5 if code.startswith("SF") else 4          # SF 多一列「结果」
    i_res = 1 if code.startswith("SF") else None
    i_src, i_song, i_art = nmeta - 3, nmeta - 2, nmeta - 1
    vcols = [(i, hdr[i]) for i in range(nmeta, iJ) if hdr[i]]   # 空表头列 = 上下半场分隔，跳过
    b1, b2 = split_blocks(rows)
    if len(b1) != len(b2):
        raise ValueError("%s 两块行数不等：%d / %d" % (sheet, len(b1), len(b2)))

    jury_total = sum(num(r[iJ]) or 0 for r in b1)
    vote_total = sum(num(r[iJ]) or 0 for r in b2)      # 下块的 Jury 列位 = 该曲总票数
    factor = jury_total / vote_total                    # 观众分 = 票数 × factor（评委分池 = 观众分池）

    voters = []
    for kind, blk in (("jury", b1), ("tele", b2)):
        for i, abbr in vcols:
            pts = {}
            for eid, r in enumerate(blk):
                v = num(r[i])
                if v:
                    pts[str(eid)] = int(v) if float(v).is_integer() else v
            if not pts:
                continue                                    # 该块未投票
            voters.append({"voter": nick(abbr), "type": kind, "points": pts})
    cast_by = set()
    for v in voters:
        cast_by.update(n.strip() for n in v["voter"].split("/"))

    entries, penalized = [], []
    for eid, (r1, r2) in enumerate(zip(b1, b2)):
        if int(num(r1[0])) != int(num(r2[0])):
            raise ValueError("%s 两块序号错位 @eid=%d" % (sheet, eid))
        nk = nick(r1[i_src])
        artist = str(r1[i_art]).strip()
        jury = num(r1[iJ]) or 0
        tele = num(r1[iT]) or 0                             # 已发布的观众分（整数），不自己重算
        raw = int(num(r2[iJ]) or 0)
        total = num(r1[iTot]) or 0
        # 选送者本场一票未投 → 总分 50%（田妈 SF1 / 杰妈 SF2）
        skip = [n.strip() for n in nk.split("/") if n.strip() not in cast_by]
        score = total / 2.0 if skip else total
        if skip:
            penalized.append(nk)
        e = {
            "eid": eid,
            "ro": int(num(r1[0])),
            "member": nk,
            "member_id": mid(nk),
            "artist": artist,
            "song": str(r1[i_song]).strip(),
            "language": meta.get(artist, {}).get("language"),
            "genre": meta.get(artist, {}).get("genre"),
            "jury_vote": jury,
            "tele_vote": tele,
            "tele_raw": raw,
            "score": score,
            "support_rate": None,
            "high_rate": None,
            "is_shadow": False,
            "rank": 0,
        }
        if i_res is not None:
            e["qualified"] = str(r1[i_res]).strip().upper() == "Q"
        if e["language"] is None:
            report.append("  ⚠ %s eid=%d 未匹配到语种/流派：%s" % (code, eid, artist))
        entries.append(e)
        # 校验：逐票和 == Jury 列；Jury + 观众分 == Total 列；观众逐票和 == 总票数
        cast = sum(num(r1[i]) or 0 for i, _ in vcols)
        if cast != jury:
            report.append("  ✗ %s ro=%s 评委逐票和 %g ≠ Jury %g" % (code, e["ro"], cast, jury))
        vcast = sum(num(r2[i]) or 0 for i, _ in vcols)
        if vcast != raw:
            report.append("  ✗ %s ro=%s 观众逐票和 %g ≠ 总票数 %g" % (code, e["ro"], vcast, raw))
        if jury + tele != total:
            report.append("  ✗ %s ro=%s Jury+Tele ≠ Total" % (code, e["ro"]))
        if abs(raw * factor - tele) > 0.52:
            report.append("  ✗ %s ro=%s 观众分 %g 与「票数×系数」%.4f 偏差过大"
                          % (code, e["ro"], tele, raw * factor))

    nj = sum(1 for v in voters if v["type"] == "jury")
    nt = sum(1 for v in voters if v["type"] == "tele")
    report.append("  %s：%d 首 / 评委 %d 人（池 %g）/ 观众 %d 人（%g 票，系数 %.6g）%s"
                  % (code, len(entries), nj, jury_total, nt, vote_total, factor,
                     ("／50% 折算：" + "、".join(penalized)) if penalized else ""))
    m = {
        "match": code,
        "venue": venue,
        "tele_mode": "votes",
        "entries": entries,
        "votes": {"voters": voters},
    }
    if penalized:
        m["note"] = ("、".join("{m:%s}" % n for n in penalized)
                     + " 未参与本场投票，其参赛作品按规则以总分的 50% 计入排名；"
                     + "表中 JURY / TELE 为折算前的原始得分。")
    return m


def build_approval(wb, sheet, code, venue, meta, report):
    """Wildcard Round / Second Chance：认可票（每人 1 票、至多 3 票）。"""
    hdr, rows = read_sheet(wb, sheet)
    iTot = hdr.index("Total")
    iPen = hdr.index("Penalty") if "Penalty" in hdr else None
    iPts = hdr.index("Pts") if "Pts" in hdr else iTot
    nmeta = 6                                              # 序号|名次|结果|来源|歌曲名称|表演者
    vcols = [(i, hdr[i]) for i in range(nmeta, iTot) if hdr[i]]
    data = [r for r in rows[1:] if num(r[0]) is not None and r[3]]

    entries = []
    for eid, r in enumerate(data):
        nk = nick(r[3])
        artist = str(r[5]).strip()
        raw = int(num(r[iTot]) or 0)
        pen = int(num(r[iPen]) or 0) if iPen is not None else 0
        pts = int(num(r[iPts]) or 0)
        e = {
            "eid": eid,
            "ro": int(num(r[0])),
            "member": nk,
            "member_id": mid(nk),
            "artist": artist,
            "song": str(r[4]).strip(),
            "language": meta.get(artist, {}).get("language"),
            "genre": meta.get(artist, {}).get("genre"),
            "jury_vote": raw,
            "tele_vote": None,
            "score": pts,
            "support_rate": None,
            "high_rate": None,
            "is_shadow": False,
            "rank": int(num(r[1])),                        # 官方名次（含并列）
            "qualified": str(r[2]).strip().upper() == "Q",
        }
        if pen:
            e["penalty"] = pen
        cast = sum(num(r[i]) or 0 for i, _ in vcols)
        if cast != raw:
            report.append("  ✗ %s ro=%s 认可票逐票和 %g ≠ Total %g" % (code, e["ro"], cast, raw))
        if raw - pen != pts:
            report.append("  ✗ %s ro=%s Total-Penalty ≠ Pts" % (code, e["ro"]))
        if e["language"] is None:
            report.append("  ⚠ %s eid=%d 未匹配到语种/流派：%s" % (code, eid, artist))
        entries.append(e)

    voters = []
    for i, abbr in vcols:
        pts = {}
        for eid, r in enumerate(data):
            v = num(r[i])
            if v:
                pts[str(eid)] = int(v)
        if not pts:
            continue
        voters.append({"voter": nick(abbr), "type": "jury", "points": pts})

    entries.sort(key=lambda e: (e["rank"], e["ro"]))   # 结果概览按名次展示（并列按出场顺序）
    npen = sum(1 for e in entries if e.get("penalty"))
    report.append("  %s：%d 首 / %d 人投票%s"
                  % (code, len(entries), len(voters), ("／%d 首被扣分" % npen) if npen else ""))
    m = {
        "match": code,
        "venue": venue,
        "format": "approval",                              # 认可票：recompute 跳过重算、stats 另计
        "entries": entries,
        "votes": {"voters": voters},
    }
    if npen:
        m["note"] = "本轮采用认可票（Approval Vote），每人至多 3 票；未投出任何一票者，其参赛作品扣 1 分。表中 PTS 为扣分后得分。"
    else:
        m["note"] = "本轮采用认可票（Approval Vote），每人至多 3 票。"
    return m


def main():
    meta, doc = load_signup_meta()
    wb = load_workbook(SRC, data_only=True)
    report = ["== 第十六届（Chongqing 2026）赛果解析 =="]
    matches = []
    for sheet, code, venue in MATCH_META:
        if code in ("WC", "SC"):
            matches.append(build_approval(wb, sheet, code, venue, meta, report))
        else:
            matches.append(build_dual(wb, sheet, code, venue, meta, report))

    # ---- 与官方 Scoreboard 交叉校验（Total / Jury / Tele / Vote）----
    ws = wb["Scoreboard"]
    rows = [r for r in ws.iter_rows(values_only=True)]
    hdr = rows[1]
    sb = [r for r in rows[2:] if num(r[0]) is not None]
    gf = {(e["member"], e["song"]): e for e in matches[4]["entries"]}
    sfe = {}
    for m in matches[:2]:
        for e in m["entries"]:
            sfe[(e["member"], e["song"])] = (m["match"], e)
    bad = 0
    for r in sb:
        key = (nick(r[1]), str(r[2]).strip())
        if num(r[4]) is not None and key in gf:                     # GF Total / Jury / Tele / Vote
            e = gf[key]
            for lbl, col, val in (("Total", 4, e["score"]), ("Jury", 6, e["jury_vote"]),
                                  ("Tele", 12, e["tele_vote"]), ("Vote", 14, e["tele_raw"])):
                if round(num(r[col])) != round(val):
                    report.append("  ✗ Scoreboard GF %s %s：官方 %g vs 计算 %g"
                                  % (key[1], lbl, num(r[col]), val)); bad += 1
        if num(r[19]) is not None and key in sfe:                   # SF Total / Jury / Tele / Vote
            _, e = sfe[key]
            for lbl, col, val in (("Total", 19, e["score"]), ("Jury", 21, e["jury_vote"]),
                                  ("Tele", 26, e["tele_vote"]), ("Vote", 27, e["tele_raw"])):
                if round(num(r[col])) != round(val):
                    report.append("  ✗ Scoreboard SF %s %s：官方 %g vs 计算 %g"
                                  % (key[1], lbl, num(r[col]), val)); bad += 1
    report.append("  Scoreboard 交叉校验：%s" % ("全部一致" if not bad else "%d 处不一致" % bad))

    # ---- members 映射（含仅投票、未选送的成员）----
    used = set()
    for m in matches:
        for e in m["entries"]:
            used.update(n.strip() for n in e["member"].split("/"))
        for v in m["votes"]["voters"]:
            used.update(n.strip() for n in v["voter"].split("/"))
    members = {n: {"id": MEMBER_ID[n], "handle": HANDLE[n]} for n in sorted(used, key=lambda x: MEMBER_ID[x])}
    report.append("  members：%d 人" % len(members))

    print("\n".join(report))
    if CHECK_ONLY:
        print("\n--check：未写文件")
        return
    doc["matches"] = matches
    doc["members"] = members
    with open(OUT, "w", encoding="utf-8", newline="\n") as fp:
        json.dump(doc, fp, ensure_ascii=False, indent=1)
    print("\n已写入 %s" % os.path.relpath(OUT, BASE))


if __name__ == "__main__":
    main()
