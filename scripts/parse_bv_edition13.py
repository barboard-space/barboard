# -*- coding: utf-8 -*-
"""Parse Barvision 第十三届 (Qiqihar 2023) → regular-13.json.

2023 重启首届（新格式，区别于 2019/2020 的 A/B/C）：三场 SF1(13)/SF2(14)/GF(18)，27 首报名、每半决赛前 9 晋级。
评委 jury = 选送成员互投；观众 public/tele = 6 名未选送成员（音/城/布/兔/T/鸽）。半决赛与决赛均拆 jury+public。
GF 折算：狼/锴 未按时投票 → 总分 50% 折算（计分板显示折前原始评委/观众分，注中说明；总分=折后）。
成员该届成绩：晋级取 GF 名次(1–18)、淘汰取总排名(19–27，两半决赛淘汰者按各自半决赛总分降序合并)。

数据源（本机磁盘）：
- 23-SF1.csv / 23-SF2.csv：半决赛干净逐票矩阵（row0 voters 从 col5 起；数据行 col0=序号 col1=选送 col2=歌手 col3=歌名 col4=总分 col5+=各票）。
- Grand Final.xlsx「总表」：决赛逐票矩阵（row0 voters cols7-32 [前 20 jury + 后 6 public]，col33=JURY col34=TELE col35=TOTAL，均原始/折前）。
- 计分表.xlsx「Sheet1」：每首 Genre。
逐票 points 用 eid 键（§2.1 契约）；public 投票人 type='tele'。
"""
import csv, json, math, os, re, sys

try:
    import openpyxl
except ImportError:
    print("need openpyxl"); sys.exit(1)

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SRC = r'D:\Genius\Barvision\Barvision 2023'
OUT = os.path.join(BASE, 'data', 'barvision', 'barvision-2023', 'regular-13.json')

PUBLIC = {'音', '城', '布', '兔', 'T', '鸽'}     # 决赛 6 名观众投票人（未选送）
FOLD = {'狼', '锴'}                               # GF 50% 折算
ALIASES = {'时': 'S'}
LANG = {
    'E': '英语', 'N': '英语', 'S': '英语', '包': '英语', '芬': '英语', '风': '英语', '汞': '英语',
    '韩': '爱沙尼亚语', '猴': '英语', '杰': '英语', '锴': '英语', '狼': '英语', '鲤': '阿尔巴尼亚语',
    '麦': '英语', '萌': '英语', '奶': '英语', '柠': '英语', '瑞': '冰岛语', '霜': '法语', '松': '英语',
    '泰': '英语', '威': '英语', '文': '法语', '星': '英语', '羊': '英语', '雨': '英语', '晕': '西班牙语',
}

NICK2ID = {}
with open(os.path.join(BASE, 'data', 'members', 'members.csv'), encoding='utf-8-sig') as f:
    for row in csv.DictReader(f):
        nk = (row.get('barboard_name') or '').strip()
        if nk:
            NICK2ID[nk] = {'id': int(row['space_id']), 'handle': (row.get('space_name') or nk).strip()}
SEEN = {}
def norm(n):
    n = (n or '').strip()
    return ALIASES.get(n, n)
def resolve(nick):
    nick = norm(nick)
    if nick in SEEN:
        return
    info = NICK2ID.get(nick + '妈') or NICK2ID.get(nick)
    SEEN[nick] = {'id': info['id'], 'handle': info['handle']} if info else {'id': None, 'handle': nick}
def mid(nick):
    return (SEEN.get(norm(nick)) or {}).get('id')
def num(v):
    if v is None or v == '':
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        m = re.match(r'\s*(-?\d+(?:\.\d+)?)', str(v))
        return float(m.group(1)) if m else None

# ── 半决赛干净 CSV ──
def read_semi_csv(path):
    rows = list(csv.reader(open(path, encoding='utf-8-sig')))
    voters = {ci: norm(c) for ci, c in enumerate(rows[0]) if c.strip() and ci >= 5}
    songs = []
    for r in rows[1:]:
        if len(r) < 5 or not r[1].strip():
            continue
        pts = {}
        for ci, v in voters.items():
            val = r[ci].strip() if ci < len(r) else ''
            if val:
                pts[v] = int(float(val))
        songs.append({'order': int(r[0]), 'sel': norm(r[1]), 'artist': r[2].strip(),
                      'song': r[3].strip(), 'total': num(r[4]), 'points': pts})
    return songs

# ── 决赛 总表 ──
def read_gf(ws, roster):
    rows = list(ws.iter_rows(values_only=True))
    voters = {ci: norm(c) for ci, c in enumerate(rows[0]) if c and 7 <= ci <= 32}
    songs = []
    for r in rows[1:]:
        sel = norm(r[2]) if len(r) > 2 else None
        if not sel or num(r[35] if len(r) > 35 else None) is None:
            continue
        pts = {ci_v: int(num(r[ci])) for ci, ci_v in voters.items() if ci < len(r) and num(r[ci]) is not None}
        jr, tr, raw = num(r[33]), num(r[34]), num(r[35])
        folded = sel in FOLD
        songs.append({
            'sel': sel, 'artist': roster[sel]['artist'], 'song': roster[sel]['song'],
            'jury': jr, 'tele': tr, 'raw_total': raw,
            'score': (math.floor(raw / 2 + 0.5) if folded else raw),
            'folded': folded, 'points': pts,
        })
    return songs

def build(songs, is_tele, genre, compute_split):
    """songs → entries[] + voters_acc。compute_split=True 时 jury/tele 由 points 求和（半决赛）。"""
    entries, vacc = [], {}
    for s in songs:
        for v in s['points']:
            vacc.setdefault(v, 'tele' if is_tele(v) else 'jury')
    # 临时名次：score 降序、并列按 tele 降序（Eurovision 平局，recompute 会重算）
    order = sorted(range(len(songs)), key=lambda i: (-(songs[i].get('score') or songs[i]['total'] or 0),
                                                     -(songs[i].get('tele') or 0)))
    rank_of = {i: r + 1 for r, i in enumerate(order)}
    for i, s in enumerate(songs):
        if compute_split:
            jv = sum(v for vn, v in s['points'].items() if not is_tele(vn))
            tv = sum(v for vn, v in s['points'].items() if is_tele(vn))
            score = s['total']
        else:
            jv, tv, score = s['jury'], s['tele'], s['score']
        entries.append({
            'eid': i, 'member': s['sel'], 'member_id': mid(s['sel']),
            'artist': s['artist'], 'song': s['song'],
            'language': LANG.get(s['sel'], ''), 'genre': genre.get(s['sel'], ''),
            'jury_vote': jv, 'tele_vote': tv, 'score': score,
            'support_rate': None, 'high_rate': None, 'is_shadow': False, 'rank': rank_of[i],
            '_pts': s['points'],
        })
    return entries, vacc

def assemble(entries, vacc):
    by = {v: {} for v in vacc}
    for e in entries:
        for v, val in e.pop('_pts').items():
            by[v][str(e['eid'])] = val
    out = []
    for v in sorted(vacc, key=lambda x: (vacc[x] == 'tele', x)):
        resolve(v)
        out.append({'voter': v, 'type': vacc[v], 'points': by[v]})
    return out

def main():
    sf1 = read_semi_csv(os.path.join(SRC, '23-SF1.csv'))
    sf2 = read_semi_csv(os.path.join(SRC, '23-SF2.csv'))
    # roster: sel → artist/song（27 首 = SF1+SF2）
    roster = {}
    for s in sf1 + sf2:
        roster.setdefault(s['sel'], {'artist': s['artist'], 'song': s['song']})
    # genre（计分表 Sheet1：col0 选送 / col3 genre）
    genre = {}
    wb = openpyxl.load_workbook(os.path.join(SRC, '计分表.xlsx'), read_only=True, data_only=True)
    for r in wb['Sheet1'].iter_rows(values_only=True):
        if r[0] and len(r) > 3 and r[3] and norm(r[0]) != 'Artist':
            genre[norm(r[0])] = str(r[3]).strip()
    wbg = openpyxl.load_workbook(os.path.join(SRC, 'Grand Final.xlsx'), read_only=True, data_only=True)
    gf = read_gf(wbg['总表'], roster)

    for nick in LANG:
        resolve(nick)

    matches = []
    for code, venue, songs in [('SF1', '半决赛一', sf1), ('SF2', '半决赛二', sf2)]:
        sels = set(s['sel'] for s in songs)               # 本场选送者 = jury，其余 = tele
        ents, vacc = build(songs, lambda v: v not in sels, genre, compute_split=True)
        for e in ents:
            e['qualified'] = e['rank'] <= 9
        matches.append({'match': code, 'venue': venue, 'entries': ents,
                        'votes': {'scale': [12,10,8,7,6,5,4,3,2,1], 'voters': assemble(ents, vacc)}})

    gents, gvacc = build(gf, lambda v: v in PUBLIC, genre, compute_split=False)
    gvoters = assemble(gents, gvacc)
    matches.append({'match': 'GF', 'venue': '总决赛',
                    'note': '狼、锴 因未按时提交投票，其选送歌曲总分按 50% 折算计入最终成绩；计分板内评委分 / 观众分为折算前原始分（狼 83 / 28、锴 36 / 18）。',
                    'entries': gents, 'votes': {'scale': [12,10,8,7,6,5,4,3,2,1], 'voters': gvoters}})

    # 总排名 1–27：GF 1–18；淘汰 9 首按半决赛总分降序 19–27
    overall = {e['member']: e['rank'] for e in gents}
    elim = sorted([(e['member'], e['score']) for m in matches[:2] for e in m['entries'] if not e['qualified']],
                  key=lambda x: -x[1])
    for i, (sel, _) in enumerate(elim):
        overall[sel] = 19 + i
    for e in gents:
        e['overall_rank'] = overall.get(e['member'])

    summary = ('Barvision Qiqihar 2023 是第 13 届 Barvision 歌曲大赛，停办两年后由雨妈发起承办、威妈包装宣传，'
               '于 2023 年在齐齐哈尔举办的重启首届。27 位成员各选送一首歌曲，经两轮半决赛筛选出 18 首晋级总决赛，'
               '由评审团与观众投票共同决出冠军。最终羊妈选送的 Best Frenz & Joywave — “Flatline” 以 159 分（评审 124 + 观众 35）'
               '夺冠并创下当时 Barvision 最高总分纪录；猴妈选送的 Raynes — “Drive You Back Home”（121 分）与'
               '风妈选送的 Haley Blais — “The Cabin”（105 分）分列亚、季军。')

    data = {
        'year': 2023, 'edition_no': 13, 'edition_name': 'Barvision Qiqihar 2023',
        'cn_name': '第十三届欧美流行歌曲个人榜吧歌曲大赛', 'version': 'regular',
        'city': '齐齐哈尔', 'host': '雨妈', 'motto': '', 'summary': summary,
        'rules': {
            'submission': '匿名参赛、私信主办方报名；每人限报 1 首并自动成为评委；发行时间 2020.1.1–2023.9.17；'
                          'lead 艺人无 BB Hot100/UK Top100、未进 Billboard Artist 100、≤1 项格莱美通类提名、≤2 首吧视常规版 Top10；'
                          '作品不得进 BB Hot100/Global200/UK 单曲榜、专辑不得进 BB200、不得进任一 BarBoard 官方榜 Top100 或评委年/半年榜。',
            'niche_standard': [
                '云村评论 ≤1000 · 艺人粉丝 ≤1万',
                'Spotify 月听众 ≤500万 · 单曲播放 ≤1000万',
                'YouTube 英语 订阅 ≤400万/MV ≤1500万 · 非英语 ≤800万/≤3000万',
            ],
            'format': '两轮半决赛各取前 9 名晋级（共 18 首）进入总决赛；评委（选送成员）+ 观众（未选送成员）投票。',
            'voting': '每位评委交 Top10 排名，按 12/10/8/7/6/5/4/3/2/1 赋分；未按时提交者其选送歌曲得分折半。',
        },
        'source': '第十三届吧视报名总则 + 23-SF1/SF2.csv（半决赛逐票）+ Grand Final 总表（决赛逐票）',
        'members': {k: SEEN[k] for k in sorted(SEEN)},
        'vote_rule': {'scale': [12,10,8,7,6,5,4,3,2,1], 'jury': '选送成员互投', 'tele': '未选送成员（观众）',
                      'note': '半决赛 jury=本场选送者、tele=其余投票人；决赛 jury 20 人 + 观众 6 人（音/城/布/兔/T/鸽）；狼/锴 50% 折算（小分为折前）。'},
        'matches': matches,
    }
    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    with open(OUT, 'w', encoding='utf-8', newline='\n') as f:
        json.dump(data, f, ensure_ascii=False, indent=1)

    print('写入', OUT)
    for m in matches:
        js = [v for v in m['votes']['voters'] if v['type'] == 'jury']
        ts = [v for v in m['votes']['voters'] if v['type'] == 'tele']
        bad = 0
        for e in m['entries']:
            jsum = sum(v['points'].get(str(e['eid']), 0) for v in m['votes']['voters'] if v['type'] == 'jury')
            tsum = sum(v['points'].get(str(e['eid']), 0) for v in m['votes']['voters'] if v['type'] == 'tele')
            if m['match'] != 'GF':  # 半决赛：jury/tele 小分和应==jury_vote/tele_vote
                if abs(jsum - e['jury_vote']) > 0.5 or abs(tsum - e['tele_vote']) > 0.5:
                    bad += 1
                    if bad <= 2: print('   ⚠', e['member'], e['song'][:16], 'J', jsum, '/', e['jury_vote'], 'T', tsum, '/', e['tele_vote'])
        print('  %s %s: %d首 jury%d tele%d%s' % (m['match'], m['venue'], len(m['entries']), len(js), len(ts), ('  ⚠%d首不符' % bad if bad else '  校验OK')))
    c = [e for e in matches[2]['entries'] if e['rank'] == 1][0]
    print('  GF冠军', c['member'], c['song'], '=', c['score'], '(J%s+T%s)' % (c['jury_vote'], c['tele_vote']))
    print('  GF 6/7:', [(e['rank'], e['member'], e['score'], 'T%s' % e['tele_vote']) for e in sorted(matches[2]['entries'], key=lambda x: x['rank'])[5:7]])
    print('  狼/锴折算:', [(e['member'], e['jury_vote'], e['tele_vote'], e['score']) for e in matches[2]['entries'] if e['member'] in FOLD])

if __name__ == '__main__':
    main()
