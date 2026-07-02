# -*- coding: utf-8 -*-
"""解析榜吧年终榜 Excel 的「总榜」sheet → data/annual/<year>.json（类 BBL 榜单）。
封面：iTunes Search API 构建期按「歌手 作品」查一次，存 600x600 URL（查不到留 null，前端 onerror 兜底）。
用法：python scripts/parse_annual_chart.py 2022 [--top 100] [--no-cover]
"""
import sys, os, json, time, re, urllib.parse, urllib.request

SRC = {
    "2023": r"D:\Genius\BarChart\吧榜文件\年终榜\2023吧榜正式版.xlsx",
    "2022": r"D:\Genius\BarChart\吧榜文件\年终榜\2022吧年榜.xlsx",
    "2021": r"D:\Genius\BarChart\吧榜文件\年终榜\2021年终榜\吧榜整理文件.xlsx",
    "2020": r"D:\Genius\BarChart\吧榜文件\年终榜\2020年终榜\00 汇总榜.xlsx",
    "2019": r"D:\Genius\BarChart\吧榜文件\年终榜\2019年终榜\2019年贴吧年终榜.xlsx",
    "2018": r"D:\Genius\BarChart\吧榜文件\年终榜\2018年终榜\000 汇总榜.xlsx",
    "2016": r"D:\Genius\BarChart\吧榜文件\年终榜\2013-2017贴吧年终榜\2016吧单曲年榜汇总.xlsx",
    "2017": r"D:\Genius\BarChart\吧榜文件\年终榜\2013-2017贴吧年终榜\2017榜吧单曲年榜.xlsx",
    "2015": r"D:\Genius\BarChart\吧榜文件\年终榜\2013-2017贴吧年终榜\个人榜吧2015年吧友汇总榜.xlsx",
    "2014": r"D:\Genius\BarChart\吧榜文件\年终榜\2013-2017贴吧年终榜\个人榜吧2014年吧友汇总榜(1).xlsx",
    "2013": r"D:\Genius\BarChart\吧榜文件\年终榜\2013-2017贴吧年终榜\13年新版吧榜 - 副本.xls",
}
# 各年源 sheet（默认「总榜」）；列名也各年不同，用 _col() 兼容
SHEET = {"2021": "吧榜豪华榜", "2020": "吧榜终版", "2023": "星号带", "2019": "Sheet1", "2018": "综合点数整理", "2016": "汇总", "2017": "Sheet1", "2015": "Sheet1", "2014": "Sheet1", "2013": "Sheet1"}
# 老式二进制 .xls（openpyxl 不支持），main() 里改用 xlrd 单独读取分支。
XLS_YEARS = {"2013"}
# 2014：唯一 sheet「Sheet1」**完全无表头**（第一行即数据，rank 直接从 1 开始）；列固定为
# rank(0)/artist(1)/-(2,分隔符无用)/title(3)/points(4,=后 21 列之和)/<21 个匿名成员列(5..25)>。
# 21 个成员列**没有任何标签**（无表头/无注释/无 defined_names），无法确定身份——用户确认该年
# 「24 榜合榜」（官方事实，与技术上能看到的 21 raw 列数不一致，只作 BOARD_COUNT 展示用，不影响管线）、
# 且**明确不做个人 Top10 / 分档助攻**（数据质量问题），但主榜与四项年度看点仍可推算。
# 见 NO_HEADER_YEARS（跳过表头读取）+ ANON_MEMBER_COLS（按固定列位置匿名统计，不产出可归因数据）。
NO_HEADER_YEARS = {"2014"}
# 2013：与 2014 不同，本身有表头（Rank/origin/Artist(s)/Title/<13 位成员>/SUM/COUNT()）——
# 用户后来发现表头本可辨认，改为**正常做身份归因**（不再匿名），见 BARE_COL_YEARS/ABBR_OVERRIDE。
# 该年官方「10 榜合榜」与技术上 13 个非空成员列不一致（同 2014 21 vs 24 的情况），BOARD_COUNT 按官方数。
# SUM/COUNT() 两列已是现成的点数/助攻数，直接走 FIXED_COLS，无需助攻数通用兜底。
ANON_MEMBER_COLS = {"2014": list(range(5, 26))}
# 2015：唯一 sheet「Sheet1」列 Rank/Artist/-/Title/(无表头,=TOTAL 重复列)/<21 位成员各占 1 列>/TOTAL，
# 官方总排名(Rank)已算好（该年为 Top 296，非常规 Top 200）、但**每位成员列存的是点数、不是名次**——
# 与 BARE_COL_YEARS 的列识别方式相同（每人仅占 1 列、表头即网名本身），故把 "2015" 也加进 BARE_COL_YEARS
# 复用列定位逻辑；但取值语义不同，交由 POINTS_NOT_RANK_YEARS 分支处理（反推各成员个人 Top N，见 main()）。
# 该年列头是英文/拼音网名（非"X妈"简称），ABBR_OVERRIDE 需为全部 21 人登记（build_abbr2id 默认按
# 昵称去"妈"匹配，网名匹配不上）。无助攻数列，改用「该行有值的成员列数」通用兜底（见 main()）。
POINTS_NOT_RANK_YEARS = {"2015", "2014", "2013"}
# 2017：唯一 sheet「Sheet1」只有 5 列（排名/艺人/歌曲/入榜数/点数），无任何成员个人列——
# 源数据本身就没有留存各大妈的详细个人提交榜单，故无法像其它年份一样从表里算出个人 Top10 /
# 分档助攻明细 / 「最多榜冠」；下方 find_member_cols 对此年份天然返回空（无列名以"排名"结尾），
# 主榜（排名/艺人/歌名/点数/入榜数=助攻数）本身完整、可正常生成。
# 「最多榜冠」改用 CHAMP_OVERRIDE 人工覆盖；各成员助攻贡献值（仅 Top100 档，来自用户提供的
# 另一份「43 榜合榜助攻贡献」统计图，非本 xlsx）另见 MANUAL_MEMBER_ASSISTS。
# 个人榜 top10 用的「全量」sheet（含所有排过的曲 → 保证满 10 条）；未定义则用主 sheet 本身。
# 2021：显示榜=豪华榜(亚洲不占位)，但 top10 从完全榜(2760 首全量)取，避免私榜曲缺失。
# 2020：显示榜=吧榜终版(300 首终版排名)，但 top10 从汇总表(2444 首全量)取，避免私榜曲缺失。
# 2023：显示榜=星号带(300 首、含 "N*" 亚洲不占位星标)，但 top10 从 Sheet1(2781 首全量) 取，避免私榜曲缺失。
# 2019：唯一 sheet「Sheet1」本身就是完整总榜（1930 首、各成员个人榜均满 top10），无需单独全量表。
# 2018：「综合点数整理」只有 Top 200（转置表，见 TRANSPOSED_YEARS），无法覆盖各成员完整 top10——
#       top10 改从 54 位成员各自的个人榜原始 Excel 读取，见 MEMBER_2018_FILES + read_2018_top10()。
FULL_SHEET = {"2021": "吧榜完全榜", "2020": "汇总表", "2023": "Sheet1"}
# 无单张全量合表、但有各大妈独立个人 sheet（sheet 名=简称，列 排名/作品/艺术家）的年份 → top10 从个人 sheet 取（最权威、满 10）
MEMBER_SHEETS = {"2022"}
# 个人榜列格式为「每人占 2 列」的年份（如 2020/2023），与默认（2021 式）单列「简称+排名」表头（如 "波排名"）不同。
# 两列顺序各年不同：2020＝(昵称整列=名次, 下一列无表头=点数)；2023＝(简称整列=点数, 下一列无表头=名次)，
# 用 PAIR_COL_RANK_SECOND 标记「名次在第二列」的年份；两种列扫描逻辑见 find_member_cols()。
PAIR_COL_YEARS = {"2020", "2023", "2016"}
PAIR_COL_RANK_SECOND = {"2023", "2016"}
# 个人榜列格式为「每人占 2 列、且两列均带完整表头」的年份（如 2019："X妈_名次"/"X妈_点数"），
# 与 PAIR_COL_YEARS（一列有表头一列无表头）不同，靠"_名次"后缀识别、忽略"_点数"列。
SUFFIX_COL_YEARS = {"2019"}
# 个人榜列格式为「每人仅占 1 列、表头即简称本身」的年份（如 2018："蛋妈"/"晕妈"…，列值直接是该成员对这首歌的名次）。
# 2015/2013 列结构相同（每人 1 列、表头即简称/网名）但列值是点数而非名次，见 POINTS_NOT_RANK_YEARS。
BARE_COL_YEARS = {"2018", "2015", "2013"}
# 源 sheet 是「转置」布局的年份（如 2018「综合点数整理」：行=字段(Rank/Artist/Title/...)+各成员，列=名次1..200）——
# main() 读取后先转置成常规「行=一首歌」布局，再走通用管线，见 main() 内 TRANSPOSED_YEARS 分支。
TRANSPOSED_YEARS = {"2018"}
# 表头过于杂乱/损坏（空字符串、纯数字 0、"#VALUE!" 公式残留等 _col() 按名字怎么也配不上）的年份，
# 直接按列下标覆盖 rank/artist/song/pts/assist（优先于 _col() 结果）。
# 2016「汇总」：col0=最终排名(已是 1..200 排好序)、col6="0 ARTIST"、col7="0SONG"、col8=纯数字 0（助攻数，
# 因是 Python 假值被 `idx={h:i for...if h}` 过滤掉，必须用下标兜底）、col9="#VALUE!"（点数，表头是残留公式错误）。
FIXED_COLS = {"2016": {"rank": 0, "artist": 6, "song": 7, "assist": 8, "pts": 9},
              "2014": {"rank": 0, "artist": 1, "song": 3, "pts": 4},
              "2013": {"rank": 0, "artist": 2, "song": 3, "pts": 17, "assist": 18}}
# 宽表个人榜区之外的元信息列名（用于从 PAIR_COL_YEARS/BARE_COL_YEARS 的表头里排除，不误当成成员列）
_META_COLS = {"终名次", "名次", "排名", "艺人", "艺术家", "歌曲", "作品", "总点数", "点数", "助攻数", "总助攻数", "前十助攻数",
              "Song", "Artist", "Artists", "In", "Points", "Rank", "Title", "Album", "Num. in Chart", "Total Points",
              # 2016「汇总」表头额外的过程性/损坏列（中间几档部分排名 + 艺人/歌曲/公式残留列，均已被 FIXED_COLS 接管，
              # 这里只是防止 find_member_cols 的 PAIR_COL_YEARS 扫描把它们误当成员列——2016 元信息列跨到第 9 列，
              # 比其它年份的 "i<4" 起扫下标更靠后）
              "最终34榜排名", "30榜排名", "26榜排名", "22榜排名", "17榜排名", "10榜排名", "0 ARTIST", "0SONG", "#VALUE!",
              # 2015 表头全大写 ARTIST/TITLE/TOTAL（与其它年份的 Title-Case 不同，需单独登记）
              "ARTIST", "TITLE", "TOTAL",
              # 2013 表头额外的结构列（origin=原始序号、Artist(s)/SUM/COUNT() 均已被 FIXED_COLS 接管）
              "origin", "Artist(s)", "SUM", "COUNT()"}
OUT_DIR = os.path.join(os.path.dirname(__file__), "..", "data", "annual")
MEMBERS_CSV = os.path.join(os.path.dirname(__file__), "..", "data", "members", "members.csv")
# 修正表（歌手/歌名规范化的唯一维护源，见该文件顶部说明）
from annual_corrections import KEEP, ABBR_OVERRIDE, ARTIST_FIX, ARTIST_RAW_FIX, SONG_FIX_EXACT, SONG_FIX_SUB, COVER_OVERRIDE, CHAMP_OVERRIDE, MANUAL_MEMBER_ASSISTS, BOARD_COUNT


# Title-Case 修复：撇号后的英文缩写/所有格被误大写（'S 'T 'M 'D 'Re 'Ve 'Ll）→ 小写
# 只改这些确定的缩写后缀，避开 O'Brien / D'Angelo / T'Aime 等名字/外语（其后非缩写词，不匹配）
_APOS_RE = re.compile(r"'(S|T|M|D|Re|Ve|Ll)\b")


def fix_text(s):
    return _APOS_RE.sub(lambda m: "'" + m.group(1).lower(), s)


_FEAT_RE = re.compile(r"\s+(?:featuring|feat|ft)\b\.?\s*", re.I)  # feat/ft/featuring → 逗号分隔
# ⚠️ 曾无 "featuring" 分支 + 无 \b 词边界，"featuring" 会被误命中「feat」子串（"uring" 残留，
# 如 "Iggy Azalea featuring Charli XCX" → "Iggy Azalea, uring Charli XCX"）；
# 2014/2015 年榜已受影响并已重新生成修复，见 CLAUDE.md #182。


def _keep_split(s):
    tmp = s
    for i, k in enumerate(KEEP):
        tmp = tmp.replace(k, "\x00%d\x00" % i)
    out = []
    for p in tmp.split(","):
        p = p.strip()
        for i, k in enumerate(KEEP):
            p = p.replace("\x00%d\x00" % i, k)
        if p:
            out.append(p)
    return out


def norm_artist(s):
    for a, b in ARTIST_RAW_FIX.items():          # 整串预修正（名字内含逗号者，拆分前先规范）
        s = s.replace(a, b)
    s = s.replace("，", ",")                      # 多艺人全角逗号分隔 → 半角（后续统一「, 」重组）
    s = _FEAT_RE.sub(", ", fix_text(s))          # feat/ft → 逗号
    parts = [ARTIST_FIX.get(p, p) for p in _keep_split(s)]  # 逐艺人官方写法
    return ", ".join(parts)                       # 统一「逗号 + 空格」


def norm_song(s):
    s = fix_text(s)
    s = SONG_FIX_EXACT.get(s, s)
    for a, b in SONG_FIX_SUB.items():
        s = s.replace(a, b)
    return s


def ckey(artist, song):
    """封面缓存键：归一（去大小写/空格/标点，保留字母数字与 CJK）→ 改格式/大小写不失效。"""
    n = lambda x: re.sub(r"\W+", "", x.lower(), flags=re.UNICODE)
    return n(artist) + "|" + n(song)


# 各大妈独立个人 sheet 表头名多变，尽力识别（识别不出则返回 []，跳过、不覆盖）
_MS_RANK = ("排名", "名次", "排位", "Rank", "RANK", "POS", "Number", "ranking", "rank")
_MS_SONG = ("作品", "歌曲", "歌名", "单曲", "曲名", "歌曲名称", "Song", "Songs", "Single", "Single Name",
            "Title", "Titles", "Track", "SONG", "SONGS", "song")
_MS_ARTIST = ("艺术家", "艺人", "歌手", "艺人名称", "Artist", "Artists", "Artist(s)", "Artist (s)",
              "Artist（s）", "Aritist", "Aritist(s)", "ARTIST", "ARITIST", "artist")


def read_member_sheet(ws):
    """从某大妈的独立个人 sheet 读其个人榜前十 [(rank, artist, song)]（表头格式多变，识别不出返回 []）。"""
    it = ws.iter_rows(values_only=True)
    try:
        hdr = next(it)
    except StopIteration:
        return []
    idx = {str(h).strip(): i for i, h in enumerate(hdr) if h is not None}
    ri = next((idx[n] for n in _MS_RANK if n in idx), None)
    si = next((idx[n] for n in _MS_SONG if n in idx), None)
    ai = next((idx[n] for n in _MS_ARTIST if n in idx), None)
    if ri is None or si is None or ai is None:
        return []
    out = []
    for row in it:
        v = row[ri]
        if isinstance(v, (int, float)) and 1 <= int(v) <= 10:
            a = norm_artist(str(row[ai] or "").strip())
            s = norm_song(str(row[si] or "").strip())
            if a and s:
                out.append((int(v), a, s))
    return out


# ── 2018：无全量/成员子表可用，个人榜 top10 改从 54 位成员各自的原始提交文件读取 ──
# 「综合点数整理」是转置后的 Top 200，无法覆盖各成员完整前十（很多人的第 1–3 名根本进不了聚合 Top 200）。
# 54 份原始文件格式各不相同（人工逐个核对列位置），故用显式 spec 表而非通用表头嗅探。
# 每条：fn=文件名（相对 2018 年终榜目录）、sheet=可选指定 sheet（默认第一个）、
# skip=跳过的表头/标题行数、a/s=艺人·歌曲列下标（0-based）；combined=单列"歌名-艺人"合并格式的列下标（如蛋妈）；
# 所有文件均已核实按名次升序排列，故不依赖显式名次列，直接用「跳过表头后的行序」当名次（第 N 个有效条目=名次 N）。
# 键 = space_id（不用简称，避免跟 ABBR_OVERRIDE 重复解析）。
MEMBER_2018_FILES = {
    17: dict(fn="01 Lee翼雨（雨妈）年终榜.xlsx", skip=0, a=1, s=2),
    133: dict(fn="02 NAVY YEAR END 100 （N妈）.xlsx", skip=1, a=2, s=1),
    191: dict(fn="03 土豆2018年榜（薯妈）.xls", skip=1, a=1, s=2),
    167: dict(fn="04 陈先森Qq 年榜TOP100（森妈）.xls", skip=0, a=0, s=1),
    157: dict(fn="05 VE Charts  2018 年榜（布妈）.xlsx", skip=0, a=0, s=1),
    195: dict(fn="06 【X妈】【Auditory】2018年个人年榜.xlsx", skip=1, a=1, s=2),
    778: dict(fn="07 农妈18单曲年榜.xlsx", skip=0, a=1, s=2, decamel=True),  # 艺人列单词间无空格，需反 camelCase
    19: dict(fn="08 lemon心语 年榜TOP100 （淋檬）.xlsx", skip=0, a=1, s=2),
    141: dict(fn="09 草妈 2018年榜.xls", skip=1, a=1, s=2),  # 文件名写「草妈」，内容实为「香草」= 草妈本人
    128: dict(fn="10 克妈-年榜.xlsx", skip=0, a=1, s=2),
    174: dict(fn="11 AzarathTrainor 2019 Year-End Chart （肥屎）.xls", skip=1, a=1, s=2),
    779: dict(fn="12 然然君.xlsx", skip=0, a=0, s=1),
    18: dict(fn="13 voiiz chart 2018年榜 (猴妈).xlsx", sheet="2018", skip=1, a=1, s=2),
    165: dict(fn="14 LORRY 年榜 （鹿妈）.xlsx", skip=1, a=2, s=1),  # 歌曲列在前、艺人列在后
    777: dict(fn="15 hc （线妈）.xlsx", skip=2, a=1, s=2),
    184: dict(fn="16 2018 Bed for you 年榜 （床妈）.xlsx", skip=1, a=2, s=3),
    138: dict(fn="17 lovezedd （Z妈）.xlsx", skip=0, a=0, s=1),
    12: dict(fn="18 WillKris Chart 2018 Year End （锴妈）.xlsx", skip=0, a=0, s=1),
    40: dict(fn="19 猹妈.xlsx", skip=0, a=1, s=2),
    176: dict(fn="20 Music Track年榜.Coke （可乐）.xlsx", sheet="Sheet2", skip=0, a=0, s=1),
    131: dict(fn="21 泰坦2018年单曲年榜 （泰妈）.xlsx", skip=2, a=1, s=2),
    119: dict(fn="22 Ocean 2018单曲年榜 （海妈）.xlsx", skip=1, a=1, s=2),
    125: dict(fn="23 converted.xlsx", skip=2, a=2, s=3),  # 原 .xls 命名表损坏 xlrd 读不出，已用 Excel 转存为 xlsx
    780: dict(fn="24 Year-end 2018 TomCYF.xlsx", sheet="TomCYF 2018", skip=1, a=0, s=1),
    781: dict(fn="25 小擦18年年榜+冠单榜.xlsx", sheet="年榜", skip=1, a=2, s=3),
    770: dict(fn="26 环球颖音2018单曲年榜 （音妈）.xlsx", skip=1, a=1, s=2),
    168: dict(fn="27 风宇凌胖2018年榜 （胖妈）.xlsx", skip=1, a=1, s=2),
    295: dict(fn="28 Song Of Freeman 2018 Year-End Charts HOT 100 （田妈）.xlsx", skip=1, a=1, s=2),
    173: dict(fn="29 YS榜2018年年榜 （俗妈）.xlsx", skip=0, a=1, s=2),
    154: dict(fn="30 小风2018年榜 （风妈）.xlsx", sheet="Feuil1", skip=1, a=2, s=3),
    123: dict(fn="31 KE - 2018 Year-End Chart (可妈).xlsx", skip=1, a=1, s=2),
    135: dict(fn="32 Jeffery's Yearly Picks 2018 Singles （瑞妈）.xlsx", skip=2, a=1, s=2),
    326: dict(fn="33 2018单曲榜年榜 （院长）.xlsx", skip=2, a=1, s=2),
    132: dict(fn="34 AA Charts 2018Year （A妈）.xlsx", skip=1, a=1, s=2),
    127: dict(fn="35 【NEOCHART】2018 Single Top 120 （苏妈）.xlsx", skip=1, a=1, s=2),
    775: dict(fn="36 SUNNYSTEWART2018单曲全年榜 （加妈）.xls", skip=1, a=2, s=3),
    20: dict(fn="37 BAg 2018年榜（榜吧版;附奖项名单）（包妈）.xlsx", skip=1, a=2, s=3),
    124: dict(fn="38 Lolo 2018年榜.xlsx", skip=1, a=1, s=2),
    156: dict(fn="39 小杰2018吧榜特供英文年榜.xls", skip=1, a=1, s=2),
    782: dict(fn="40 2018 JD HOT 100年榜.xlsx", skip=0, a=0, s=1),
    776: dict(fn="41 冰妈年榜.xls", skip=0, a=0, s=1),
    164: dict(fn="42 liangliang年榜.xlsx", skip=0, a=1, s=2),
    158: dict(fn="43 漫画小鸣2018年榜.xlsx", skip=1, a=1, s=2),
    172: dict(fn="44 2018清妈年榜-1.xlsx", sheet="工作表 1 - 2018 Oliver Year-End Ch", skip=2, a=1, s=2),
    130: dict(fn="45 麦妈Book1.xlsx", skip=0, a=1, s=2),
    7: dict(fn="46 GENIUS TOP 50 2018年榜.xlsx", skip=1, a=1, s=2),
    117: dict(fn="47 晕妈的2018单曲年榜.xlsx", skip=1, a=1, s=2),
    118: dict(fn="48 RAZ Chart 2018年榜（参与吧榜版本）.xlsx", sheet="正式榜单", skip=0, a=2, s=3),
    153: dict(fn="49 DJL 2018单曲榜（筛选后）.xlsx", skip=1, a=2, s=3),
    129: dict(fn="50 火火2018-年榜.xlsx", skip=1, a=1, s=2),
    120: dict(fn="51 城妈2018年榜.xlsx", skip=1, a=2, s=3),
    189: dict(fn="52 麻雀2018年榜.xlsx", skip=1, a=1, s=2),
    185: dict(fn="53 2018年年榜_Clean.xlsx", sheet="年榜榜单", skip=1, combined=3),  # 单列"歌名-艺人"合并格式
    190: dict(fn="54 5D Single Chart 2018年榜top100 - 副本.xlsx", skip=2, a=1, s=2),
}
_DECAMEL_RE = re.compile(r"(?<=[a-z])(?=[A-Z])")


def _fmt_cell(v):
    """单元格转文本：整数值的 float（如 2002.0，Excel 把纯数字歌名存成数值）去掉多余 .0。"""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v or "")


def _read_2018_personal_chart(spec, base_dir):
    """按 spec 读一位成员的原始个人榜文件，返回 [(rank, artist, song)]（最多前十，行序即名次）。"""
    path = os.path.join(base_dir, spec["fn"])
    if path.lower().endswith(".xls"):
        import xlrd
        wb = xlrd.open_workbook(path)
        ws = wb.sheet_by_name(spec["sheet"]) if spec.get("sheet") else wb.sheet_by_index(0)
        all_rows = [ws.row_values(r) for r in range(ws.nrows)]
    else:
        import openpyxl
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        ws = wb[spec["sheet"]] if spec.get("sheet") else wb[wb.sheetnames[0]]
        all_rows = list(ws.iter_rows(values_only=True))
    data_rows = all_rows[spec.get("skip", 0):]
    out = []
    for row in data_rows:
        if len(out) >= 10:
            break
        if "combined" in spec:
            raw = row[spec["combined"]] if spec["combined"] < len(row) else None
            if not raw or "-" not in str(raw):
                continue
            song, artist = str(raw).split("-", 1)
        else:
            ai, si = spec["a"], spec["s"]
            if ai >= len(row) or si >= len(row):
                continue
            artist, song = row[ai], row[si]
        artist = _fmt_cell(artist).replace("\xa0", " ").strip()
        song = _fmt_cell(song).replace("\xa0", " ").strip()
        if spec.get("decamel"):
            artist = _DECAMEL_RE.sub(" ", artist)
        if not artist or not song:
            continue
        out.append((len(out) + 1, norm_artist(artist), norm_song(song)))
    return out


def find_member_cols(header, year):
    """从表头识别「个人榜简称 → 名次列下标」。四种格式：
    - 默认（如 2021）：单列，表头「简称+排名」（如 "波排名"，可带 "求和项:" 前缀）。
    - PAIR_COL_YEARS（如 2020/2023）：每人占 2 列，表头列非空、紧跟一列无表头；
      名次在第一列还是第二列由 PAIR_COL_RANK_SECOND 决定（2020＝第一列，2023＝第二列）。
    - SUFFIX_COL_YEARS（如 2019）：每人占 2 列，两列均带完整表头「X妈_名次」「X妈_点数」，
      靠 "_名次" 后缀识别、直接取该列为名次列，"_点数" 列忽略。
    - BARE_COL_YEARS（如 2018，转置后）：每人仅占 1 列，表头即简称本身（如 "蛋妈"），列值直接是名次。
    返回 (abbrs, col_of_abbr)，abbrs 已去除「妈」后缀（供 build_abbr2id 用统一规则匹配 members.csv；
    本身已是单字简称的年份如 2023 不受影响，末字不是「妈」不会被误切）。
    """
    col_of = {}
    if year in SUFFIX_COL_YEARS:
        for i, h in enumerate(header):
            if not h:
                continue
            hs = str(h).strip()
            if not hs.endswith("_名次"):
                continue
            name = hs[:-3]
            a = name[:-1] if name.endswith("妈") else name
            col_of[a] = i
    elif year in PAIR_COL_YEARS:
        rank_second = year in PAIR_COL_RANK_SECOND
        for i, h in enumerate(header):
            if i < 4 or not h:
                continue
            hs = str(h).strip()
            if hs in _META_COLS:
                continue
            a = hs[:-1] if hs.endswith("妈") else hs
            col_of[a] = i + 1 if rank_second else i
    elif year in BARE_COL_YEARS:
        for i, h in enumerate(header):
            if not h:
                continue
            hs = str(h).strip()
            if hs in _META_COLS:
                continue
            a = hs[:-1] if hs.endswith("妈") else hs
            col_of[a] = i
    else:
        for i, h in enumerate(header):
            if i >= 5 and h and str(h).endswith("排名"):
                a = str(h).replace("求和项:", "").replace("排名", "")
                col_of[a] = i
    return list(col_of.keys()), col_of


def build_abbr2id(abbrs, year=None):
    """把总榜里的个人榜简称（松/威/N/时…）映射到 space_id；ABBR_OVERRIDE 按年消歧。"""
    import csv as _csv
    mem = []
    with open(MEMBERS_CSV, encoding="utf-8-sig") as f:
        rd = _csv.reader(f)
        next(rd)
        for r in rd:
            if r and r[0]:
                mem.append((int(r[0]), r[1]))
    ov = ABBR_OVERRIDE.get(str(year), {})
    out, miss = {}, []
    for a in abbrs:
        if a in ov:
            out[a] = ov[a]
            continue
        c = [sid for sid, name in mem if name == a + "妈" or name == a or (name and len(a) == 1 and name[0] == a)]
        if len(c) == 1:
            out[a] = c[0]
        else:
            miss.append(a)
    if miss:
        sys.stderr.write("  ! 简称未唯一匹配 %d 个: %s\n" % (len(miss), miss))
    return out


_SPLIT_RE = re.compile(r"\s*(?:,|/|&|、|×|\bfeat\.?|\bft\.?|\bwith\b|\bx\b)\s*", re.I)
_PAREN_RE = re.compile(r"\s*[\(（\[][^\)）\]]*[\)）\]]")


def _primary_artist(artist):
    """取第一位主唱（feat/ft/with/,/&/、 之前）。"""
    parts = _SPLIT_RE.split(artist)
    return (parts[0] if parts else artist).strip() or artist


def _strip_paren(song):
    """去掉歌名里的括号补充（(feat. X) / (From The Series...) 等）。"""
    return _PAREN_RE.sub("", song).strip() or song


def _norm(s):
    """归一：仅保留字母数字与 CJK，用于歌名相近校验。"""
    return re.sub(r"[^a-z0-9一-鿿]+", "", (s or "").lower())


def _artwork(r):
    a = r.get("artworkUrl100") or r.get("artworkUrl60") or r.get("artworkUrl30")
    if not a:
        return None
    return re.sub(r"/\d+x\d+bb\.", "/600x600bb.", a)


def _fetch_itunes(term, limit=5):
    """查 iTunes（song 实体）；网络错误退避重试一次；每次真实请求后礼貌 sleep。"""
    url = "https://itunes.apple.com/search?term=%s&entity=song&limit=%d" % (
        urllib.parse.quote(term[:180]), limit)
    for attempt in range(2):
        try:
            req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
            d = json.load(urllib.request.urlopen(req, timeout=15))
            time.sleep(0.4)
            return d.get("results", [])
        except Exception as e:
            if attempt == 0:
                time.sleep(2.5)  # 疑似限流，退避后重试
                continue
            sys.stderr.write("  ! cover fetch fail [%s]: %s\n" % (term[:60], e))
            time.sleep(0.4)
    return []


def itunes_cover(artist, song, cache):
    key = ckey(artist, song)
    if (artist, song) in COVER_OVERRIDE:
        cache[key] = COVER_OVERRIDE[(artist, song)]
        return cache[key]
    if key in cache:
        return cache[key]
    pa, sp = _primary_artist(artist), _strip_paren(song)
    # 多级回退：完整艺人+歌名 → 主艺人+歌名 → 主艺人+去括号歌名
    terms = []
    for a in (artist, pa):
        for s in (song, sp):
            t = a + " " + s
            if t not in terms:
                terms.append(t)
    art = None
    for term in terms:
        for r in _fetch_itunes(term):
            art = _artwork(r)
            if art:
                break
        if art:
            break
    # 末级回退：仅歌名，但校验返回曲名与目标相近，避免误配到同名别曲
    if not art:
        ns, nsp = _norm(song), _norm(sp)
        for r in _fetch_itunes(song, limit=8):
            tn = _norm(r.get("trackName", ""))
            if tn and (ns in tn or tn in ns or (nsp and nsp in tn)):
                art = _artwork(r)
                if art:
                    break
    cache[key] = art
    return art


def main():
    year = sys.argv[1] if len(sys.argv) > 1 else "2022"
    top = 100
    nocover = "--no-cover" in sys.argv
    if "--top" in sys.argv:
        top = int(sys.argv[sys.argv.index("--top") + 1])
    wb = None
    if year in XLS_YEARS:
        # 老式 .xls 二进制格式，openpyxl 不支持，改用 xlrd；空单元格 xlrd 给 ''，统一转 None
        # 与 openpyxl 语义对齐，下游 isinstance/is None 判断无需为此年份特殊处理。
        import xlrd
        xwb = xlrd.open_workbook(SRC[year])
        xws = xwb.sheet_by_name(SHEET.get(year, "Sheet1"))
        _n = lambda v: None if v == "" else v
        all_data_rows = [tuple(_n(v) for v in xws.row_values(r)) for r in range(xws.nrows)]
        header = all_data_rows[0]
        rows = iter(all_data_rows[1:])
    else:
        import openpyxl
        wb = openpyxl.load_workbook(SRC[year], read_only=True, data_only=True)
        ws = wb[SHEET.get(year, "总榜")]
        if year in TRANSPOSED_YEARS:
            # 源 sheet 是转置布局（行=字段名/成员，列=名次1..N）——先转置成常规「行=一首歌」布局
            # 再走通用管线：转置后第 0 行 = 原第 0 列 = ['Rank','Artist','Title',...,'蛋妈',...,'X妈']（新表头），
            # 之后每行 = 一首歌（各字段 + 每位成员对这首歌的名次）
            matrix = list(ws.iter_rows(values_only=True))
            transposed = list(zip(*matrix))
            rows = iter(transposed[1:])
            header = transposed[0]
        elif year in NO_HEADER_YEARS:
            # 完全无表头，第一行即数据——不消费首行，列全靠 FIXED_COLS 按下标定位
            rows = ws.iter_rows(values_only=True)
            header = []
        else:
            rows = ws.iter_rows(values_only=True)
            header = next(rows)
    # 定位列（各年列名不同：艺术家/艺人、作品/歌曲、点数/总点数、助攻数/总助攻数）
    idx = {h: i for i, h in enumerate(header) if h}

    def _col(*names):
        for n in names:
            if n in idx:
                return idx[n]
        return None
    ci = {"rank": _col("排名", "名次", "终名次", "Rank"), "artist": _col("艺术家", "艺人", "Artists", "Artist", "ARTIST"), "song": _col("作品", "歌曲", "Song", "Track", "Title", "TITLE"),
          "pts": _col("点数", "总点数", "Points", "Total Points", "TOTAL"), "assist": _col("助攻数", "总助攻数", "In", "Charts", "Num. in Chart", "入榜数")}
    if ci["rank"] is None:
        ci["rank"] = 0  # 排名列无表头（如 2023，第 1 列即排名，无列名）
    if year in FIXED_COLS:
        ci.update(FIXED_COLS[year])  # 表头过烂，直接按列下标覆盖（见 FIXED_COLS 注释）
    # 成员个人榜列 → col_index -> space_id（两种表头格式，见 find_member_cols）；
    # ANON_MEMBER_COLS 年份无表头可用于身份识别，rank_cols 用「列下标=伪 id」仅供匿名聚合
    # （助攻数/最多榜冠计数），伪 id 不写入任何输出、也不产出 member-annual-index.json 数据。
    if year in ANON_MEMBER_COLS:
        rank_cols = {c: c for c in ANON_MEMBER_COLS[year]}
    else:
        abbrs, col_of_abbr = find_member_cols(header, year)
        abbr2id = build_abbr2id(abbrs, year)
        rank_cols = {col: abbr2id[a] for a, col in col_of_abbr.items() if a in abbr2id}
    # 复用已有 JSON 里「已命中」的封面（不缓存 null）→ 每次重跑只重试缺封面项、不动已命中的
    fp = os.path.join(OUT_DIR, year + ".json")
    cache = {}
    if os.path.exists(fp):
        try:
            old = json.load(open(fp, encoding="utf-8"))
            for oe in old.get("entries", []):
                if oe.get("cover"):
                    cache[ckey(oe["artist"], oe["song"])] = oe["cover"]
        except Exception:
            pass
    # 也复用成员年榜索引里已抓到的 top10 封面（含 Top200 之外的私榜曲），避免重复请求
    midx_path = os.path.join(OUT_DIR, "member-annual-index.json")
    if os.path.exists(midx_path):
        try:
            mold = json.load(open(midx_path, encoding="utf-8"))
            for ydata in mold.values():
                for mv in ydata.values():
                    for te in mv.get("top10", []):
                        if te.get("cover"):
                            cache[ckey(te["artist"], te["song"])] = te["cover"]
        except Exception:
            pass
    # 名次：整数=欧美占位曲；字符串带星（如 34* / 145**）=亚洲单曲（华语/K-POP），不占位、只插入展示
    STAR_RE = re.compile(r"^\s*(\d+)\s*(\**)")
    all_rows = list(rows)
    # 成员列存点数、非名次的年份（如 2015）：反推每位成员的个人排名——
    # 用该成员在全表（不限于官方 Top N）里的点数降序排位，代替其它年份直接读列值=名次的做法。
    personal_rank = {}  # sid -> {(artist,song): 反推名次}
    personal_top1 = {}  # sid -> (artist,song)，反推第 1 名（供 no1_by 判定用）
    if year in POINTS_NOT_RANK_YEARS:
        personal_raw = {sid: [] for sid in set(rank_cols.values())}
        for row in all_rows:
            a = norm_artist(str(row[ci["artist"]] or "").strip())
            s = norm_song(str(row[ci["song"]] or "").strip())
            if not a or not s:
                continue
            for col, sid in rank_cols.items():
                v = row[col]
                if isinstance(v, (int, float)) and v > 0:
                    personal_raw[sid].append((v, a, s))
        for sid, lst in personal_raw.items():
            lst.sort(key=lambda x: -x[0])
            personal_rank[sid] = {(a, s): i + 1 for i, (v, a, s) in enumerate(lst)}
            if lst:
                personal_top1[sid] = (lst[0][1], lst[0][2])
    entries = []
    for row in all_rows:
        rk_raw = row[ci["rank"]]
        star = 0
        if isinstance(rk_raw, (int, float)):
            rk = int(rk_raw)
        elif isinstance(rk_raw, str):
            m = STAR_RE.match(rk_raw)
            if not m:
                continue
            rk = int(m.group(1))
            star = len(m.group(2))
        else:
            continue
        if rk < 1 or rk > top:
            continue
        artist = norm_artist(str(row[ci["artist"]] or "").strip())
        song = norm_song(str(row[ci["song"]] or "").strip())
        if not artist or not song:
            continue
        # 助攻数：优先读显式列；无此列但有成员个人列（rank_cols）时，兜底按「该行有值的成员列数」统计
        # （某成员是否包含这首歌，与该列存的是名次还是点数无关——只看是否非空）
        if ci["assist"] is not None and row[ci["assist"]] is not None:
            assists = int(row[ci["assist"]])
        elif rank_cols:
            assists = sum(1 for col in rank_cols if row[col] is not None)
        else:
            assists = None
        e = {"rank": rk, "artist": artist, "song": song,
             "points": round(float(row[ci["pts"]] or 0), 1),
             "assists": assists,
             "cover": None}
        if star:
            e["star"] = star
        # 个人榜冠军数：多少位大妈把这首歌排在其个人榜第 1（no1_by = 这些大妈的 space_id）
        if year in POINTS_NOT_RANK_YEARS:
            no1_by = sorted(sid for sid, top1 in personal_top1.items() if top1 == (artist, song))
        else:
            no1_by = sorted(sid for col, sid in rank_cols.items()
                            if isinstance(row[col], (int, float)) and int(row[col]) == 1)
        if no1_by:
            e["no1"] = len(no1_by)
            if year not in ANON_MEMBER_COLS:  # 伪 id 不是真实 space_id，不能对外暴露
                e["no1_by"] = no1_by
        # 「最多榜冠」人工覆盖（源数据缺个人榜逐曲名次、无法自动统计的年份，见 CHAMP_OVERRIDE）
        co = CHAMP_OVERRIDE.get(year)
        if co and co["artist"] == artist and co["song"] == song:
            e["no1"] = co["no1"]
        if not nocover:
            e["cover"] = itunes_cover(artist, song, cache)
            sys.stderr.write("  %3d%s. %s — %s  %s\n" % (rk, "*" * star, artist, song, "✓" if e["cover"] else "—"))
        entries.append(e)
    # 同名次下多首不占位曲的星号数重新计算：源数据里的星号数是原表格顺序，不一定按点数——
    # 按点数降序重新分配 1,2,3...（点数越高星号越少），与下面的显示排序（也按点数降序）保持一致
    by_rank = {}
    for e in entries:
        if e.get("star"):
            by_rank.setdefault(e["rank"], []).append(e)
    for group in by_rank.values():
        group.sort(key=lambda x: -x["points"])
        for i, e in enumerate(group):
            e["star"] = i + 1
    # 同名次（并排名次相同）时按点数降序排——真实占位曲点数必然更高，天然排最前；
    # 多首不占位曲并排同一名次时，点数高的排前面（而非按星号数排，星号数只表示"隔了几首不占位曲"，与排序先后无关）
    entries.sort(key=lambda x: (x["rank"], -x["points"]))
    out = {"year": int(year), "title": "%s 榜吧年终榜" % year, "count": len(entries), "entries": entries}
    if year in BOARD_COUNT:
        out["boards"] = BOARD_COUNT[year]
    os.makedirs(OUT_DIR, exist_ok=True)
    with open(fp, "w", encoding="utf-8") as f:
        json.dump(out, f, ensure_ascii=False, indent=1)
    covered = sum(1 for e in entries if e["cover"])
    print("写出 %s | %d 条 | 封面命中 %d/%d" % (fp, len(entries), covered, len(entries)))

    # ── 成员年榜聚合 → data/annual/member-annual-index.json（个人主页「个人年榜」板块）──
    TIERS = [10, 20, 50, 100, 200]
    # ANON_MEMBER_COLS 年份：伪 id 不对应真实成员，个人 Top10/分档助攻不产出（用户确认，数据质量问题）——
    # sids 置空即可让下方 assist/top10/magg 全部自然产出空结果，personal_rank（供 no1 匿名判定用）不受影响
    sids = [] if year in ANON_MEMBER_COLS else sorted(set(rank_cols.values()))
    assist = {s: {t: 0 for t in TIERS} for s in sids}     # 助攻分档（占位/欧美曲）：按年榜名次累计（前 N）
    assist_sh = {s: {t: 0 for t in TIERS} for s in sids}  # 助攻分档（亚洲不占位曲）：单独统计、括号标注
    if sids and year in POINTS_NOT_RANK_YEARS:
        # 用反推名次（personal_rank，见上）分档，而非直接读列值——该年无亚洲不占位曲概念，assist_sh 全 0
        for sid, rankmap in personal_rank.items():
            for (a, s), rk in rankmap.items():
                for t in TIERS:
                    if rk <= t:
                        assist[sid][t] += 1
    elif sids:
        for row in all_rows:                                  # 助攻分档 = 主显示榜（豪华榜）
            rk_raw = row[ci["rank"]]
            star = 0
            if isinstance(rk_raw, (int, float)):
                rk = int(rk_raw)
            elif isinstance(rk_raw, str):
                mm = STAR_RE.match(rk_raw)
                if not mm:
                    continue
                rk = int(mm.group(1))
                star = len(mm.group(2))
            else:
                continue
            if rk < 1:
                continue
            for col, sid in rank_cols.items():
                if not isinstance(row[col], (int, float)) or row[col] <= 0:
                    continue  # <=0 排除「未上榜」哨兵值（如 2018 用 0 表示未排入，而非 None）
                tgt = assist if star == 0 else assist_sh      # 占位曲计主数 / 亚洲不占位曲计括号
                for t in TIERS:
                    if rk <= t:
                        tgt[sid][t] += 1

    # 个人榜前十：优先从「完全榜」全量 sheet 取（保证满 10 条）；否则用主 sheet
    def collect_top10(rows_, ci_, rcols_):
        t = {s: [] for s in sids}
        for row in rows_:
            a = norm_artist(str(row[ci_["artist"]] or "").strip())
            s = norm_song(str(row[ci_["song"]] or "").strip())
            if not a or not s:
                continue
            for col, sid in rcols_.items():
                v = row[col]
                if isinstance(v, (int, float)) and 1 <= int(v) <= 10:
                    t[sid].append((int(v), a, s))
        return t
    if year in POINTS_NOT_RANK_YEARS:
        # 反推排名（personal_rank）本身已是完整"点数降序"排位，直接取前十即可，无需额外全量表
        top10 = {sid: [(rk, a, s) for (a, s), rk in personal_rank.get(sid, {}).items() if rk <= 10] for sid in sids}
    elif year in FULL_SHEET:
        fws = wb[FULL_SHEET[year]]
        fit = fws.iter_rows(values_only=True)
        fh = next(fit)
        fidx = {h: i for i, h in enumerate(fh) if h}
        fci = {"artist": next((fidx[n] for n in ("艺术家", "艺人", "Artists") if n in fidx), None),
               "song": next((fidx[n] for n in ("作品", "歌曲", "Song") if n in fidx), None)}
        _, fcol_of_abbr = find_member_cols(fh, year)
        frcols = {col: abbr2id[a] for a, col in fcol_of_abbr.items() if a in abbr2id}
        top10 = collect_top10(list(fit), fci, frcols)
    elif year == "2018":
        # 「综合点数整理」只有 Top 200、覆盖不了各成员完整前十，改从各自原始提交文件读取（见 MEMBER_2018_FILES）
        base_dir = os.path.dirname(SRC[year])
        top10 = {sid: [] for sid in sids}
        for sid in sids:
            spec = MEMBER_2018_FILES.get(sid)
            if not spec:
                continue
            try:
                top10[sid] = _read_2018_personal_chart(spec, base_dir)
            except Exception as e:
                sys.stderr.write("  ! 2018 个人榜读取失败 sid=%s (%s): %s\n" % (sid, spec.get("fn"), e))
    else:
        top10 = collect_top10(all_rows, ci, rank_cols)
    # 补全：MEMBER_SHEETS 年份里，总榜不足 10 条的成员，从其独立个人 sheet 尽力补齐（个人 sheet 表头多变）
    if year in MEMBER_SHEETS:
        for abbr, sid in abbr2id.items():
            if len(top10.get(sid, [])) >= 10 or abbr not in wb.sheetnames:
                continue
            st = read_member_sheet(wb[abbr])
            if len(st) > len(top10.get(sid, [])):
                top10[sid] = st
    magg = {}
    for sid in sids:
        lst = sorted(top10[sid], key=lambda x: x[0])[:10]
        if not lst and not any(assist[sid].values()) and not any(assist_sh[sid].values()):
            continue
        t10 = []
        for r, a, s in lst:
            t10.append({"rank": r, "artist": a, "song": s,
                        "cover": (None if nocover else itunes_cover(a, s, cache))})
        m = {"assists": {"t%d" % t: assist[sid][t] for t in TIERS}, "top10": t10}
        if any(assist_sh[sid].values()):
            m["assists_shadow"] = {"t%d" % t: assist_sh[sid][t] for t in TIERS}
        magg[str(sid)] = m
    # 人工覆盖：源数据缺各成员详细个人榜的年份（如 2017），无法从主榜算出分档助攻/Top10，
    # 改用用户提供的另一份统计人工登记（仅 t100 有值，其余档位/Top10 留空，前端显示"—"）
    for sid, t100 in MANUAL_MEMBER_ASSISTS.get(year, {}).items():
        magg[str(sid)] = {
            "assists": {"t%d" % t: (t100 if t == 100 else None) for t in TIERS},
            "top10": [], "no_detail": True,
        }
    idx_path = os.path.join(OUT_DIR, "member-annual-index.json")
    idx = {}
    if os.path.exists(idx_path):
        try:
            idx = json.load(open(idx_path, encoding="utf-8"))
        except Exception:
            idx = {}
    idx[year] = magg
    with open(idx_path, "w", encoding="utf-8") as f:
        json.dump(idx, f, ensure_ascii=False, indent=1)
    print("成员年榜聚合 → %s | %d 位大妈" % (idx_path, len(magg)))


if __name__ == "__main__":
    main()
